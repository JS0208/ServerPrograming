from flask import Flask, render_template, request, redirect, url_for, flash # flash 임포트
from markupsafe import Markup # Markup을 markupsafe에서 가져옵니다.
from db_config import get_db_connection
import mysql.connector
import FinanceDataReader as fdr
import numpy as np
import pandas_ta as ta
import pandas as pd
import math # // 연산자 대신 정확한 내림을 위해 math.floor 사용 가능, 또는 //로도 충분
import json # JSON 처리를 위해 추가
from datetime import datetime, timedelta
from openpyxl import Workbook
#from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils.dataframe import dataframe_to_rows # Pandas DataFrame을 Excel 행으로 변환
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from flask import send_file # 파일 전송을 위해 추가
import io # 메모리 내 파일 처리를 위해 추가

from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user

from flask import make_response # make_response 추가
from weasyprint import HTML, CSS # WeasyPrint 임포트
#from weasyprint.fonts import FontConfiguration # 폰트 설정을 위해 (선택적)
import os # 파일 경로 처리를 위해
import tempfile # 임시 파일 생성을 위해 (차트 이미지 등)
import matplotlib
matplotlib.use('Agg') # 백엔드에서 Matplotlib 사용 설정 (중요)
import matplotlib.pyplot as plt
from matplotlib.dates import AutoDateLocator, DateFormatter # 날짜 포맷팅
import base64 # 이미지를 base64로 인코딩하여 HTML에 직접 삽입하기 위함

STRATEGY_TYPES = [
    "선택안함", # 기본값 또는 빈 값 허용 시
    "추세추종 (Trend Following)",
    "평균회귀 (Mean Reversion)",
    "모멘텀 (Momentum)",
    "변동성 돌파 (Volatility Breakout)",
    "가치 투자 (Value Investing)",
    "기타 (Custom/Other)"
]

app = Flask(__name__)
app.secret_key = 'your_very_secret_key_here' # 플래시 메시지를 위한 시크릿 키 설정 (실제 운영시에는 복잡하고 안전한 값으로 변경)

# Flask-Login 설정
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login' # 로그인 안된 사용자가 @login_required 페이지 접근 시 리다이렉트될 엔드포인트
login_manager.login_message = "로그인이 필요한 페이지입니다."
login_manager.login_message_category = "warning"

COMMISSION_RATE = 0.00015  # 0.015%
SELL_TAX_RATE = 0.0015     # 0.15% (2025년 기준 예시)

class User(UserMixin):
    def __init__(self, id, username):
        self.id = id
        self.username = username

def generate_chart_for_pdf(dates, data, title, ylabel, filename_suffix, is_drawdown=False):
    if not dates or not data:
        return None
        
    fig, ax = plt.subplots(figsize=(10, 5)) # PDF에 적합한 크기로 조절
    
    # 날짜 데이터가 문자열이면 datetime 객체로 변환 (이미 변환되어 있다고 가정)
    # dates_dt = [datetime.strptime(d, '%Y-%m-%d') for d in dates]
    dates_dt = pd.to_datetime(dates) # pandas DatetimeIndex로 가정

    ax.plot(dates_dt, data, color='blue' if not is_drawdown else 'red', linewidth=1.5)
    ax.set_title(title, fontsize=14)
    ax.set_xlabel("날짜", fontsize=10)
    ax.set_ylabel(ylabel, fontsize=10)
    ax.grid(True, linestyle='--', alpha=0.7)
    
    # X축 날짜 포맷 설정
    locator = AutoDateLocator()
    formatter = DateFormatter('%Y-%m-%d') # 또는 '%y-%m' 등 데이터 범위에 맞게
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(formatter)
    fig.autofmt_xdate() # 날짜 라벨이 겹치지 않도록 자동 조절

    plt.tight_layout() # 레이아웃 조절

    # 이미지를 base64 문자열로 변환
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=100) # 해상도 조절 가능
    img_buffer.seek(0)
    img_base64 = base64.b64encode(img_buffer.read()).decode('utf-8')
    plt.close(fig) # 메모리 해제
    
    return f"data:image/png;base64,{img_base64}"

@login_manager.user_loader
def load_user(user_id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn is None:
            return None
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, username FROM users WHERE id = %s", (user_id,))
        user_data = cursor.fetchone()
        if user_data:
            return User(id=user_data['id'], username=user_data['username'])
        return None
    except Exception as e:
        print(f"Error loading user: {e}")
        return None
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
        
# run_backtest_logic 함수 내에서 사용할 안전한 포맷팅 헬퍼 함수
def format_kpi_value(value, precision=2, is_percentage=True, is_currency=False, default_na_str="N/A"):
    """
    KPI 값을 안전하게 포맷팅합니다.
    None이면 default_na_str을 반환합니다.
    숫자면 지정된 정밀도로 포맷팅하고, 필요시 '%' 또는 ' 원'을 붙입니다.
    """
    if value is None:
        return default_na_str
    
    if isinstance(value, str) and "N/A" in value: # 이미 "N/A" 처리된 문자열
        return value

    try:
        num_value = float(value) # 문자열 형태의 숫자도 처리 시도
        if is_currency:
            return f"{num_value:,.0f} 원"
        elif is_percentage:
            return f"{num_value:.{precision}f}%"
        else: # 일반 숫자
            return f"{num_value:.{precision}f}"
    except (ValueError, TypeError):
        # float 변환 실패 시 또는 기타 타입이면 원본 문자열 또는 default_na_str 반환
        return str(value) if isinstance(value, str) else default_na_str
    
# 백테스트 실행 로직을 별도 함수로 분리 (재활용 위함)
def run_backtest_logic(ticker, start_date_str, end_date_str, conditions_from_db, 
                       initial_capital, position_sizing_method, 
                       fixed_amount_param, fixed_percentage_param, 
                       commission_rate_const, sell_tax_rate_const):
    # --- 데이터 로드 ---
    stock_data_df = fdr.DataReader(ticker, start_date_str, end_date_str)
    if stock_data_df.empty or len(stock_data_df) < 2:
        raise ValueError(f"{ticker} ({start_date_str}~{end_date_str}) 데이터 부족")
    if not isinstance(stock_data_df.index, pd.DatetimeIndex):
        stock_data_df.index = pd.to_datetime(stock_data_df.index)

    # --- 지표 계산 ---
    indicator_columns = ['Close']
    df_for_ta = stock_data_df.copy()
    df_for_ta.rename(columns={'Open':'open', 'High':'high', 'Low':'low', 'Close':'close', 'Volume':'volume'}, inplace=True)
    
    # sma_col_name_generated는 현재 코드에서 직접 사용되지 않으므로 제거해도 무방합니다.
    # indicator_columns 리스트로 관리됩니다.

    for condition in conditions_from_db:
        indicator_type = condition['indicator_type'].upper()
        params_str = condition['value']
        try:
            if indicator_type == 'SMA':
                period = int(params_str); col_name = f'SMA_{period}'
                df_for_ta.ta.sma(length=period, append=True, col_names=(col_name,))
                if col_name not in indicator_columns: indicator_columns.append(col_name)
            elif indicator_type == 'RSI':
                parts = [p.strip() for p in params_str.split(',')]; period = int(parts[0])
                col_name = f'RSI_{period}'
                df_for_ta.ta.rsi(length=period, append=True, col_names=(col_name,))
                if col_name not in indicator_columns: indicator_columns.append(col_name)
            elif indicator_type == 'MACD':
                parts = [p.strip() for p in params_str.split(',')]; fast = int(parts[0]); slow = int(parts[1]); signal_p = int(parts[2])
                macd_df = df_for_ta.ta.macd(fast=fast, slow=slow, signal=signal_p, append=False)
                macd_line_col = f'MACD_{fast}_{slow}'; signal_line_col = f'MACDsignal_{fast}_{slow}_{signal_p}'
                df_for_ta[macd_line_col] = macd_df[f'MACD_{fast}_{slow}_{signal_p}']
                df_for_ta[signal_line_col] = macd_df[f'MACDs_{fast}_{slow}_{signal_p}']
                if macd_line_col not in indicator_columns: indicator_columns.append(macd_line_col)
                if signal_line_col not in indicator_columns: indicator_columns.append(signal_line_col)
        except Exception as e:
            raise ValueError(f"지표 '{indicator_type}' 계산 오류 (값: {params_str}): {e}")

    for col in df_for_ta.columns:
        if col not in stock_data_df.columns and col not in ['open', 'high', 'low', 'close', 'volume']:
            stock_data_df[col] = df_for_ta[col]
    stock_data_df.dropna(inplace=True) # 모든 지표 계산 후 NaN이 있는 행 제거
    if stock_data_df.empty:
        raise ValueError("모든 지표 계산 후 유효한 데이터가 없습니다.")

    # --- 개별 조건 신호 ---
    buy_cols_to_check = []
    sell_cols_to_check = []
    for idx, condition in enumerate(conditions_from_db):
        indicator_type = condition['indicator_type'].upper(); params_str = condition['value']
        buy_col = f'buy_cond_{idx}'; sell_col = f'sell_cond_{idx}'
        stock_data_df[buy_col] = False; stock_data_df[sell_col] = False
        # ... (이전과 동일한 조건별 신호 생성 로직) ...
        if indicator_type == 'SMA':
            period = int(params_str); sma_col = f'SMA_{period}'
            if sma_col in stock_data_df.columns:
                stock_data_df[buy_col] = stock_data_df['Close'] > stock_data_df[sma_col]
                stock_data_df[sell_col] = stock_data_df['Close'] < stock_data_df[sma_col]
        elif indicator_type == 'RSI':
            parts = [p.strip() for p in params_str.split(',')]; period = int(parts[0]); oversold = int(parts[1]); overbought = int(parts[2])
            rsi_col = f'RSI_{period}'
            if rsi_col in stock_data_df.columns:
                stock_data_df[buy_col] = stock_data_df[rsi_col] < oversold
                stock_data_df[sell_col] = stock_data_df[rsi_col] > overbought
        elif indicator_type == 'MACD':
            parts = [p.strip() for p in params_str.split(',')]; fast = int(parts[0]); slow = int(parts[1]); signal_p = int(parts[2])
            macd_line_col = f'MACD_{fast}_{slow}'; signal_line_col = f'MACDsignal_{fast}_{slow}_{signal_p}'
            if macd_line_col in stock_data_df.columns and signal_line_col in stock_data_df.columns:
                stock_data_df[buy_col] = stock_data_df[macd_line_col] > stock_data_df[signal_line_col]
                stock_data_df[sell_col] = stock_data_df[macd_line_col] < stock_data_df[signal_line_col]
        if buy_col not in buy_cols_to_check: buy_cols_to_check.append(buy_col)
        if sell_col not in sell_cols_to_check: sell_cols_to_check.append(sell_col)

    # --- 최종 신호 ---
    # buy_cols_to_check 등이 비어있을 경우를 대비하여 default 값 설정
    final_buy_signal = stock_data_df[buy_cols_to_check].all(axis=1) if buy_cols_to_check else pd.Series([False]*len(stock_data_df), index=stock_data_df.index)
    final_sell_signal = stock_data_df[sell_cols_to_check].all(axis=1) if sell_cols_to_check else pd.Series([False]*len(stock_data_df), index=stock_data_df.index)
    stock_data_df['Signal'] = 0 
    stock_data_df.loc[final_buy_signal, 'Signal'] = 1
    stock_data_df.loc[final_sell_signal, 'Signal'] = -1
    
    # --- 포트폴리오 시뮬레이션 ---
    # ... (이전과 동일한 포트폴리오 시뮬레이션 로직) ...
    cash = initial_capital; shares = 0; portfolio_value = initial_capital; position = 0
    stock_data_df['Portfolio_Value'] = initial_capital
    trades_count = 0; winning_trades_count = 0; last_buy_price = 0.0
    for i in range(len(stock_data_df)):
        current_signal = stock_data_df['Signal'].iloc[i]; current_price = stock_data_df['Close'].iloc[i]
        if position == 1: current_portfolio_value = shares * current_price + cash
        else: current_portfolio_value = cash
        stock_data_df.loc[stock_data_df.index[i], 'Portfolio_Value'] = current_portfolio_value
        if current_signal == 1 and position == 0:
            target_investment_cash = 0
            if position_sizing_method == 'all_in': target_investment_cash = cash 
            elif position_sizing_method == 'fixed_amount': target_investment_cash = min(fixed_amount_param if fixed_amount_param is not None else 0, cash) # None 체크
            elif position_sizing_method == 'fixed_percentage': target_investment_cash = min(current_portfolio_value * (fixed_percentage_param if fixed_percentage_param is not None else 0), cash) # None 체크
            effective_buy_price_per_share = current_price * (1 + commission_rate_const)
            shares_can_buy = math.floor(target_investment_cash / effective_buy_price_per_share) if effective_buy_price_per_share > 0 and target_investment_cash > 0 else 0
            if shares_can_buy > 0:
                buy_cost_principal = shares_can_buy * current_price; commission_cost = buy_cost_principal * commission_rate_const
                if cash >= buy_cost_principal + commission_cost:
                    shares = shares_can_buy; cash -= (buy_cost_principal + commission_cost); position = 1
                    trades_count += 1; last_buy_price = current_price
        elif current_signal == -1 and position == 1:
            sell_value_total = shares * current_price; commission_cost = sell_value_total * commission_rate_const
            tax_cost = sell_value_total * sell_tax_rate_const; net_proceeds = sell_value_total - commission_cost - tax_cost
            cash += net_proceeds
            if current_price > last_buy_price : winning_trades_count +=1
            shares = 0; position = 0
    if position == 1:
        sell_value_total = shares * stock_data_df['Close'].iloc[-1]; commission_cost = sell_value_total * commission_rate_const
        tax_cost = sell_value_total * sell_tax_rate_const; cash += (sell_value_total - commission_cost - tax_cost)
    final_portfolio_value = cash if cash is not None else initial_capital # final_portfolio_value가 None이 되지 않도록

    # --- 성과 지표 ---
    # 계산 전 None 가능성 있는 값들 기본값 처리
    initial_capital_val = initial_capital if initial_capital is not None else 0.0
    final_portfolio_value_val = final_portfolio_value if final_portfolio_value is not None else initial_capital_val

    total_return_pct = ((final_portfolio_value_val - initial_capital_val) / initial_capital_val) * 100 if initial_capital_val > 0 else 0.0
    
    stock_data_df['Peak'] = stock_data_df['Portfolio_Value'].cummax()
    stock_data_df['Drawdown'] = (stock_data_df['Portfolio_Value'] - stock_data_df['Peak']) / stock_data_df['Peak']
    # max_drawdown_pct 계산 시 stock_data_df['Drawdown']이 비어있거나 모두 NaN인 경우 0.0으로 처리
    max_drawdown_pct = stock_data_df['Drawdown'].min() * 100 if not stock_data_df['Drawdown'].empty and not stock_data_df['Drawdown'].isnull().all() else 0.0

    win_rate_pct = (winning_trades_count / trades_count) * 100 if trades_count > 0 else 0.0
    
    cagr_raw = 0.0 # 숫자형 CAGR을 저장할 변수
    if not stock_data_df.empty:
        actual_start_date_dt = stock_data_df.index[0].to_pydatetime()
        actual_end_date_dt = stock_data_df.index[-1].to_pydatetime()
        years = (actual_end_date_dt - actual_start_date_dt).days / 365.25
    else: # stock_data_df가 비어있는 경우 (예: 모든 데이터가 NaN으로 제거됨)
        start_dt_obj = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_dt_obj = datetime.strptime(end_date_str, '%Y-%m-%d')
        years = (end_dt_obj - start_dt_obj).days / 365.25

    cagr_display_str = "N/A" # 최종 문자열 표시용
    if years > 0.0027: # 대략 1일 이상일 경우 (0으로 나누는 것 방지 및 의미 있는 기간)
        if initial_capital_val > 0 and final_portfolio_value_val > 0: # 초기값, 최종값 모두 양수일 때만 의미있는 CAGR
            cagr_raw = (((final_portfolio_value_val / initial_capital_val) ** (1 / years)) - 1) * 100
            cagr_display_str = format_kpi_value(cagr_raw, is_percentage=True)
        elif final_portfolio_value_val == initial_capital_val: # 변동 없음
            cagr_raw = 0.0
            cagr_display_str = format_kpi_value(cagr_raw, is_percentage=True)
        # 그 외 (손실 등)는 cagr_raw는 0.0으로 두고, cagr_display_str는 "N/A" 유지 (혹은 손실 CAGR 표시)
    elif total_return_pct != 0: # 기간이 매우 짧지만 수익/손실이 있는 경우
        cagr_display_str = "N/A (기간 부족)"
    else: # 기간도 짧고 수익/손실도 없는 경우
        cagr_raw = 0.0
        cagr_display_str = format_kpi_value(cagr_raw, is_percentage=True)


    # --- 결과 데이터 준비 ---
    # 포지션 사이징 정보 문자열 생성 시 None 체크 강화
    ps_fixed_amount_str = format_kpi_value(fixed_amount_param, is_percentage=False, is_currency=True, default_na_str='0 원') if fixed_amount_param is not None else '미사용'
    ps_fixed_percentage_str = format_kpi_value(fixed_percentage_param * 100 if fixed_percentage_param is not None else None, is_percentage=True, default_na_str='0.00%') if fixed_percentage_param is not None else '미사용'
    
    position_sizing_info_str = {
        "all_in": "전량 투자", 
        "fixed_amount": f"고정 금액 ({ps_fixed_amount_str})",
        "fixed_percentage": f"고정 비율 ({ps_fixed_percentage_str})"
    }.get(position_sizing_method, "알 수 없음")

    result_summary_kpis = {
        "total_return_pct": format_kpi_value(total_return_pct, is_percentage=True),
        "max_drawdown_pct": format_kpi_value(max_drawdown_pct, is_percentage=True),
        "win_rate_pct": f"{format_kpi_value(win_rate_pct, is_percentage=True)} ({winning_trades_count}/{trades_count})",
        "cagr_pct": cagr_display_str, # 이미 문자열로 처리됨
        "num_trades": trades_count, # 정수이므로 그대로 사용
        "initial_capital": format_kpi_value(initial_capital, is_percentage=False, is_currency=True),
        "final_portfolio_value": format_kpi_value(final_portfolio_value, is_percentage=False, is_currency=True),
        "applied_costs_info": f"매수/매도 수수료: {format_kpi_value(commission_rate_const*100, precision=3, is_percentage=True)}, 매도세: {format_kpi_value(sell_tax_rate_const*100, precision=3, is_percentage=True)}",
        "position_sizing_info": position_sizing_info_str,
        # kpis 딕셔너리에 DB 저장용 원시 숫자 값도 포함시킬 수 있음 (선택 사항)
        # 예: "cagr_raw": cagr_raw if isinstance(cagr_raw, (int, float)) else None 
    }
    
    # ... (차트 데이터 준비 및 display_cols 설정은 이전과 동일) ...
    chart_labels = [date.strftime('%Y-%m-%d') for date in stock_data_df.index]
    chart_portfolio_values = stock_data_df['Portfolio_Value'].round(2).tolist()
    chart_drawdown_values = stock_data_df['Drawdown'].round(4).tolist()
    monthly_portfolio_values = stock_data_df['Portfolio_Value'].resample('M').last()
    monthly_returns = monthly_portfolio_values.pct_change().fillna(0)
    histogram_frequencies = []; histogram_labels = []
    if not monthly_returns.empty and len(monthly_returns) > 1:
        hist, bin_edges = np.histogram(monthly_returns * 100, bins=10) # 수익률을 %로 변환하여 히스토그램
        histogram_frequencies = hist.tolist()
        for j in range(len(bin_edges) - 1): histogram_labels.append(f"{bin_edges[j]:.1f}%~{bin_edges[j+1]:.1f}%")
    
    # display_cols에 indicator_columns 외 필요한 컬럼 추가
    display_cols_set = set(indicator_columns)
    for col_list in [buy_cols_to_check, sell_cols_to_check]: # buy_cols_to_check 등은 실제 컬럼명 리스트여야 함
        for col_name in col_list:
            if col_name in stock_data_df.columns:
                display_cols_set.add(col_name)
    for col_name in ['Signal', 'Portfolio_Value', 'Drawdown']:
        if col_name in stock_data_df.columns:
            display_cols_set.add(col_name)
    
    final_display_columns = list(display_cols_set)
    if 'Close' in final_display_columns: # Close 컬럼을 맨 앞으로
        final_display_columns.insert(0, final_display_columns.pop(final_display_columns.index('Close')))
    else: # Close가 없으면 추가 (데이터 무결성 확인 필요)
        final_display_columns.insert(0, 'Close')


    return result_summary_kpis, stock_data_df, final_display_columns, chart_labels, chart_portfolio_values, chart_drawdown_values, histogram_labels, histogram_frequencies

# SP_Project/app.py 에 새로 추가할 함수

def portfolio_backtest_logic(tickers, weights, start_date_str, end_date_str,
                             conditions_from_db, initial_portfolio_capital,
                             position_sizing_method_portfolio, # 포트폴리오 레벨에서의 포지션 사이징 (예: 자산별 초기 할당)
                             fixed_amount_param_portfolio, 
                             fixed_percentage_param_portfolio,
                             commission_rate_const, sell_tax_rate_const,
                             strategy_name_display):
    """
    포트폴리오 백테스트를 실행하는 함수.
    단순화를 위해, 초기 자본을 각 티커에 비중대로 분배하고,
    각 티커는 독립적으로 run_backtest_logic (단일 종목 백테스트 로직)을 따르되,
    자금은 각자 할당된 초기 자본 내에서 운영한다고 가정. (매우 단순화된 모델)
    더 정교한 모델은 포트폴리오 전체 현금을 공유하고, 리밸런싱 등을 고려해야 함.
    """
    all_assets_data = {}
    min_len = float('inf')
    common_index = None

    # 1. 모든 티커 데이터 로드 및 공통 기간 설정
    for ticker in tickers:
        try:
            df = fdr.DataReader(ticker, start_date_str, end_date_str)
            if df.empty or len(df) < 2:
                flash(f"{ticker} 데이터를 가져오거나 분석하기에 충분하지 않습니다.", "warning")
                # return None # 또는 해당 티커 제외하고 진행
                continue # 이 티커는 건너뜀
            if not isinstance(df.index, pd.DatetimeIndex):
                df.index = pd.to_datetime(df.index)
            
            # 지표 계산 (run_backtest_logic의 지표 계산 부분 재활용 또는 별도 함수화)
            # 여기서는 단순화를 위해 지표계산은 각 run_backtest_logic 호출 시 수행하도록 함
            # 또는 여기서 일괄적으로 모든 데이터에 대해 지표를 미리 계산할 수도 있음
            
            all_assets_data[ticker] = df
            if common_index is None:
                common_index = df.index
            else:
                common_index = common_index.intersection(df.index)
            min_len = min(min_len, len(df))
        except Exception as e:
            flash(f"{ticker} 데이터 로드 중 오류: {str(e)}", "warning")
            # return None
            continue # 이 티커는 건너뜀
    
    if not all_assets_data or common_index is None or len(common_index) < 2:
        flash("유효한 데이터를 가진 자산이 없거나 공통 거래일이 부족합니다.", "danger")
        return None

    # 모든 데이터를 공통 인덱스로 정렬 및 필터링
    for ticker in list(all_assets_data.keys()): # list()로 감싸서 반복 중 삭제 가능하게
        if ticker in all_assets_data:
            all_assets_data[ticker] = all_assets_data[ticker].reindex(common_index).sort_index()
            # NaN이 너무 많은 경우 해당 자산 제외 고려
            if all_assets_data[ticker]['Close'].isnull().sum() > len(all_assets_data[ticker]) * 0.5:
                flash(f"{ticker}는 공통 기간 동안 데이터가 너무 많이 누락되어 제외합니다.", "warning")
                del all_assets_data[ticker]
                # weights도 재조정 필요 (여기서는 단순화를 위해 생략)

    if not all_assets_data:
        flash("모든 자산이 데이터 부족으로 제외되었습니다.", "danger")
        return None
        
    # weights와 all_assets_data의 ticker 목록 일치시키기 (제외된 티커 반영)
    # 이 부분은 더 정교한 처리가 필요합니다. 여기서는 간단히 weights의 길이가 같다면 진행.
    if len(weights) != len(all_assets_data):
        flash("데이터 로드 후 티커 수와 비중 수가 일치하지 않습니다. 비중을 재조정하거나 티커를 확인해주세요.", "warning")
        # 간단히 동일 비중으로 재할당 (더 나은 방법: 사용자에게 알리고 중단)
        active_tickers = list(all_assets_data.keys())
        weights = [1.0 / len(active_tickers)] * len(active_tickers)
        tickers = active_tickers # tickers 리스트도 업데이트
    
    # 2. 자산별 백테스트 실행 및 포트폴리오 가치 계산 준비
    portfolio_daily_values = pd.DataFrame(index=common_index)
    portfolio_daily_values['Portfolio_Value'] = 0.0
    
    total_trades = 0
    total_winning_trades = 0

    # 각 자산에 할당된 초기 자본금
    asset_initial_capitals = {ticker: initial_portfolio_capital * weight for ticker, weight in zip(tickers, weights)}

    for i, ticker in enumerate(tickers):
        if ticker not in all_assets_data:
            continue
            
        asset_data_df = all_assets_data[ticker].copy()
        asset_initial_capital = asset_initial_capitals[ticker]

        try:
            # 단일 종목 백테스트 로직(run_backtest_logic)을 각 자산에 적용
            # 주의: run_backtest_logic은 현재 Flask 요청 컨텍스트에 의존하지 않도록 수정되어야 함
            #      또는 필요한 부분만 추출하여 새로운 함수로 만들어야 함.
            #      여기서는 run_backtest_logic이 순수 계산 함수라고 가정.
            #      포지션 사이징 관련 파라미터는 자산별로 할당된 자본 내에서 동작하도록 조정.
            
            # run_backtest_logic을 호출하기 위한 준비
            # 포지션 사이징: 'all_in'은 해당 자산에 할당된 자본 내에서 전부 투자
            # 'fixed_amount'는 해당 자산에 대해 지정된 고정 금액 (asset_initial_capital을 넘지 않도록)
            # 'fixed_percentage'는 해당 자산에 할당된 자본의 특정 비율
            
            # 현재 run_backtest_logic은 flash 메시지 등을 포함할 수 있어 직접 재활용이 어려울 수 있습니다.
            # 핵심 계산 로직만 분리한 함수 (예: calculate_single_asset_backtest)를 사용하는 것이 좋습니다.
            # 여기서는 개념적으로 run_backtest_logic을 호출한다고 가정하고,
            # 그 결과 중 포트폴리오 가치 변화(일별)를 가져온다고 가정합니다.

            # ---- 임시: 실제로는 아래 부분을 정교한 단일 에셋 백테스트 함수로 대체해야 함 ----
            # 이 예제에서는 각 자산이 독립적으로 거래되고, 그 결과를 합산하는 방식으로 단순화합니다.
            # run_backtest_logic을 직접 호출하는 대신, 그 내부 로직을 참고하여 포트폴리오용으로 재구성.
            
            # 예시: 단일 자산의 가치 변화만 가져오는 함수가 있다고 가정
            # asset_kpis, asset_df_processed, _, _, asset_portfolio_values_ts, _, _, _ = \
            #     run_backtest_logic(ticker, start_date_str, end_date_str, conditions_from_db, 
            #                        asset_initial_capital, position_sizing_method_portfolio, 
            #                        fixed_amount_param_portfolio, fixed_percentage_param_portfolio, # 이 값들은 자산별로 적용
            #                        commission_rate_const, sell_tax_rate_const)
            # 
            # if asset_portfolio_values_ts: # Series 형태라고 가정 (index=날짜, value=자산가치)
            #     portfolio_daily_values['Portfolio_Value'] += pd.Series(asset_portfolio_values_ts, index=common_index).fillna(method='ffill').fillna(asset_initial_capital) # NaN 처리 중요
            # else: # 백테스트 실패 시 초기 자본으로 채움
            #     portfolio_daily_values['Portfolio_Value'] += asset_initial_capital

            # --- 좀 더 현실적인 접근: run_backtest_logic의 핵심을 가져와서 사용 ---
            # 아래는 run_backtest_logic의 시뮬레이션 부분을 단순화하여 적용한 예시입니다.
            # 실제로는 지표 계산, 신호 생성 등이 각 asset_data_df에 대해 수행되어야 합니다.
            
            # (1) 각 자산별 지표 계산 (기존 run_backtest_logic의 지표 계산 부분)
            df_for_ta = asset_data_df.copy()
            df_for_ta.rename(columns={'Open':'open', 'High':'high', 'Low':'low', 'Close':'close', 'Volume':'volume'}, inplace=True)
            # indicator_columns = ['close'] # pandas_ta는 소문자 컬럼명을 사용하므로, rename 후 사용
            
            for condition in conditions_from_db:
                indicator_type = condition['indicator_type'].upper()
                params_str = condition['value']
                try:
                    if indicator_type == 'SMA':
                        period = int(params_str)
                        col_name = f'SMA_{period}'
                        # df_for_ta.ta.sma(length=period, append=True, col_names=(col_name,)) # 원본은 df_for_ta에 추가
                        asset_data_df[col_name] = ta.sma(asset_data_df['Close'], length=period) # 직접 asset_data_df에 추가
                    elif indicator_type == 'RSI':
                        parts = [p.strip() for p in params_str.split(',')]
                        period = int(parts[0])
                        # oversold, overbought는 신호 생성 시 사용되므로 여기서는 RSI 값만 계산
                        col_name = f'RSI_{period}'
                        asset_data_df[col_name] = ta.rsi(asset_data_df['Close'], length=period)
                    elif indicator_type == 'MACD':
                        parts = [p.strip() for p in params_str.split(',')]
                        fast, slow, signal_p = int(parts[0]), int(parts[1]), int(parts[2])
                        macd_df = ta.macd(asset_data_df['Close'], fast=fast, slow=slow, signal=signal_p)
                        # MACD 결과 컬럼명이 ta 버전에 따라 다를 수 있음. 예: MACD_12_26_9, MACDs_12_26_9, MACDh_12_26_9
                        # 실제 컬럼명을 확인하고 asset_data_df에 할당해야 합니다.
                        asset_data_df[f'MACD_{fast}_{slow}_{signal_p}'] = macd_df[f'MACD_{fast}_{slow}_{signal_p}']
                        asset_data_df[f'MACDsignal_{fast}_{slow}_{signal_p}'] = macd_df[f'MACDs_{fast}_{slow}_{signal_p}']
                        asset_data_df[f'MACDhist_{fast}_{slow}_{signal_p}'] = macd_df[f'MACDh_{fast}_{slow}_{signal_p}']
                    # 필요한 다른 지표들도 여기에 추가
                except Exception as e:
                    print(f"Error calculating indicator {indicator_type} for {ticker}: {e}") # 어떤 티커, 어떤 지표에서 문제인지 확인
                    # 해당 지표 계산 실패 시, 이 자산의 백테스트를 건너뛸지, 아니면 해당 조건 없이 진행할지 결정 필요
                    # asset_data_df[col_name] = pd.NA # 또는 적절한 오류 처리
                    continue # 일단 이 조건 계산은 건너뛰고 다음 조건으로

            # 지표 계산 후 NaN 값 처리 (중요!)
            asset_data_df.dropna(inplace=True) # 지표 계산으로 인해 앞부분에 NaN이 생김
            if asset_data_df.empty:
                print(f"{ticker} has no data after indicator calculation and dropna.")
                continue # 이 자산은 더 이상 진행 불가

            # (2) 각 자산별 신호 생성 (기존 run_backtest_logic의 신호 생성 부분)
            # buy_cols_to_check 와 sell_cols_to_check는 각 자산마다 초기화되어야 합니다.
            asset_buy_conditions = []
            asset_sell_conditions = []

            for cond_idx, condition_spec in enumerate(conditions_from_db): # conditions_from_db 사용
                indicator_type = condition_spec['indicator_type'].upper()
                params_str = condition_spec['value']
                
                # 각 조건에 따른 boolean Series 생성
                current_buy_signal = pd.Series([False]*len(asset_data_df), index=asset_data_df.index)
                current_sell_signal = pd.Series([False]*len(asset_data_df), index=asset_data_df.index)

                if indicator_type == 'SMA':
                    period = int(params_str)
                    sma_col = f'SMA_{period}'
                    if sma_col in asset_data_df.columns:
                        current_buy_signal = asset_data_df['Close'] > asset_data_df[sma_col]
                        current_sell_signal = asset_data_df['Close'] < asset_data_df[sma_col]
                elif indicator_type == 'RSI':
                    parts = [p.strip() for p in params_str.split(',')]
                    period = int(parts[0])
                    # RSI 조건에는 보통 기준선(들)이 포함됨. 예: "14,30,70" (기간, 과매도선, 과매수선)
                    # 여기서는 params_str이 "기간,과매도기준,과매수기준" 형태라고 가정. 예: "14,30,70"
                    if len(parts) == 3:
                        oversold = int(parts[1])
                        overbought = int(parts[2])
                        rsi_col = f'RSI_{period}'
                        if rsi_col in asset_data_df.columns:
                            current_buy_signal = asset_data_df[rsi_col] < oversold
                            current_sell_signal = asset_data_df[rsi_col] > overbought
                    else:
                        print(f"RSI params for {ticker} are not in 'period,oversold,overbought' format: {params_str}")
                elif indicator_type == 'MACD':
                    parts = [p.strip() for p in params_str.split(',')]
                    fast, slow, signal_p = int(parts[0]), int(parts[1]), int(parts[2])
                    macd_line_col = f'MACD_{fast}_{slow}_{signal_p}'
                    signal_line_col = f'MACDsignal_{fast}_{slow}_{signal_p}'
                    if macd_line_col in asset_data_df.columns and signal_line_col in asset_data_df.columns:
                        current_buy_signal = asset_data_df[macd_line_col] > asset_data_df[signal_line_col]
                        current_sell_signal = asset_data_df[macd_line_col] < asset_data_df[signal_line_col]
                
                asset_buy_conditions.append(current_buy_signal)
                asset_sell_conditions.append(current_sell_signal)

            # 최종 신호: 모든 매수 조건 만족 시 매수, 모든 매도 조건 만족 시 매도
            # 주의: 현재 로직은 모든 조건이 AND로 연결됩니다.
            if asset_buy_conditions: # 매수 조건이 하나라도 있다면
                final_buy_signal = pd.concat(asset_buy_conditions, axis=1).all(axis=1)
            else: # 매수 조건이 없다면 항상 False
                final_buy_signal = pd.Series([False]*len(asset_data_df), index=asset_data_df.index)

            if asset_sell_conditions: # 매도 조건이 하나라도 있다면
                final_sell_signal = pd.concat(asset_sell_conditions, axis=1).all(axis=1)
            else: # 매도 조건이 없다면 항상 False
                final_sell_signal = pd.Series([False]*len(asset_data_df), index=asset_data_df.index)

            asset_data_df['Signal'] = 0 
            asset_data_df.loc[final_buy_signal, 'Signal'] = 1
            # 매수 신호와 매도 신호가 동시에 발생할 경우 매수 우선 또는 특정 규칙 필요. 여기서는 매도 신호가 덮어쓸 수 있음.
            # 좀 더 명확히 하려면: asset_data_df.loc[final_sell_signal & (asset_data_df['Signal'] == 0), 'Signal'] = -1
            asset_data_df.loc[final_sell_signal, 'Signal'] = -1 
            
            # (3) 각 자산별 포트폴리오 시뮬레이션 (자산별 할당된 자본으로)
            cash = asset_initial_capital; shares = 0; position = 0;
            asset_values_over_time = pd.Series(index=asset_data_df.index, dtype=float)
            asset_trades = 0; asset_winning_trades = 0; last_buy_price = 0

            for date_idx in range(len(asset_data_df)):
                current_signal = asset_data_df['Signal'].iloc[date_idx]
                current_price = asset_data_df['Close'].iloc[date_idx]
                
                current_asset_value = cash + (shares * current_price if position == 1 else 0)
                asset_values_over_time.iloc[date_idx] = current_asset_value

                if current_signal == 1 and position == 0: # Buy
                    target_investment_cash = 0
                    # 포지션 사이징 (자산별 할당된 자본 내에서)
                    if position_sizing_method_portfolio == 'all_in': target_investment_cash = cash
                    elif position_sizing_method_portfolio == 'fixed_amount': target_investment_cash = min(fixed_amount_param_portfolio, cash)
                    # fixed_percentage는 asset_initial_capital의 % 또는 현재 cash의 % 등 정의 필요. 여기서는 cash의 %로 가정
                    elif position_sizing_method_portfolio == 'fixed_percentage': target_investment_cash = min(cash * fixed_percentage_param_portfolio, cash)

                    effective_buy_price_per_share = current_price * (1 + commission_rate_const)
                    shares_can_buy = math.floor(target_investment_cash / effective_buy_price_per_share) if effective_buy_price_per_share > 0 and target_investment_cash > 0 else 0
                    if shares_can_buy > 0:
                        buy_cost_principal = shares_can_buy * current_price; commission = buy_cost_principal * commission_rate_const
                        if cash >= buy_cost_principal + commission:
                            shares = shares_can_buy; cash -= (buy_cost_principal + commission); position = 1
                            asset_trades += 1; last_buy_price = current_price
                elif current_signal == -1 and position == 1: # Sell
                    sell_value_total = shares * current_price; commission = sell_value_total * commission_rate_const
                    tax = sell_value_total * sell_tax_rate_const; net_proceeds = sell_value_total - commission - tax
                    cash += net_proceeds
                    if current_price > last_buy_price: asset_winning_trades +=1
                    shares = 0; position = 0
            
            if position == 1: # 마지막 날 강제 청산
                sell_value_total = shares * asset_data_df['Close'].iloc[-1]; commission = sell_value_total * commission_rate_const
                tax = sell_value_total * sell_tax_rate_const; cash += (sell_value_total - commission - tax)
                asset_values_over_time.iloc[-1] = cash # 최종 가치 업데이트
            
            # asset_values_over_time을 common_index에 맞춰서 portfolio_daily_values에 더함
            # reindex 후 ffill을 통해 주말 등 누락된 날짜의 가치를 이전 값으로 채움
            # 초기값은 해당 자산의 asset_initial_capital로 채움
            portfolio_daily_values['Portfolio_Value'] += asset_values_over_time.reindex(common_index).fillna(method='ffill').fillna(value=asset_initial_capital)

            total_trades += asset_trades
            total_winning_trades += asset_winning_trades

        except Exception as e:
            flash(f"{ticker} 백테스트 중 오류: {str(e)}", "danger")
            # 오류 발생 시 해당 자산은 초기 가치로만 더해지도록 처리 (또는 전체 중단)
            portfolio_daily_values['Portfolio_Value'] += asset_initial_capitals.get(ticker, 0) # 오류난 자산은 초기값으로만 기여
            print(f"Error during backtest for {ticker}: {e}")
            import traceback
            traceback.print_exc()


    # 3. 포트폴리오 전체 성과 지표 계산
    final_portfolio_value = portfolio_daily_values['Portfolio_Value'].iloc[-1]
    total_return_pct = ((final_portfolio_value - initial_portfolio_capital) / initial_portfolio_capital) * 100 if initial_portfolio_capital > 0 else 0.0
    
    portfolio_daily_values['Peak'] = portfolio_daily_values['Portfolio_Value'].cummax()
    portfolio_daily_values['Drawdown'] = (portfolio_daily_values['Portfolio_Value'] - portfolio_daily_values['Peak']) / portfolio_daily_values['Peak']
    max_drawdown_pct = portfolio_daily_values['Drawdown'].min() * 100 if not portfolio_daily_values['Drawdown'].empty and not portfolio_daily_values['Drawdown'].isnull().all() else 0.0
    win_rate_pct = (total_winning_trades / total_trades) * 100 if total_trades > 0 else 0.0

    # CAGR 계산 (기존 run_backtest_logic 참고)
    actual_start_date_dt = portfolio_daily_values.index[0].to_pydatetime()
    actual_end_date_dt = portfolio_daily_values.index[-1].to_pydatetime()
    years = (actual_end_date_dt - actual_start_date_dt).days / 365.25
    cagr_display_str = "N/A"
    if years > 0.0027:
        if initial_portfolio_capital > 0 and final_portfolio_value > 0:
            cagr_raw = (((final_portfolio_value / initial_portfolio_capital) ** (1 / years)) - 1) * 100
            cagr_display_str = format_kpi_value(cagr_raw, is_percentage=True)
        # ... (기타 CAGR 조건 처리) ...
    elif total_return_pct != 0: cagr_display_str = "N/A (기간 부족)"
    else: cagr_display_str = format_kpi_value(0.0, is_percentage=True)

    # 포지션 사이징 정보 문자열 (포트폴리오 레벨)
    # 이 부분은 포트폴리오에서 포지션 사이징이 어떻게 정의되느냐에 따라 달라집니다.
    # 여기서는 "자산별 초기 할당 후, 각 자산 내에서 ..." 의 의미로 해석.
    # 좀 더 명확한 포트폴리오 레벨의 포지션 사이징 설명이 필요합니다.
    position_sizing_info_str = f"자산별 초기 비중 분배 후, {position_sizing_method_portfolio} 적용"
    # ... (fixed_amount, fixed_percentage 값에 대한 설명 추가)

    kpis = {
        "strategy_name": f"{strategy_name_display} (포트폴리오)",
        "ticker": ", ".join(tickers), # 모든 티커 나열
        "period": f"{start_date_str} ~ {end_date_str}",
        "initial_capital": format_kpi_value(initial_portfolio_capital, is_percentage=False, is_currency=True),
        "final_portfolio_value": format_kpi_value(final_portfolio_value, is_percentage=False, is_currency=True),
        "total_return_pct": format_kpi_value(total_return_pct, is_percentage=True),
        "cagr_pct": cagr_display_str,
        "max_drawdown_pct": format_kpi_value(max_drawdown_pct, is_percentage=True),
        "win_rate_pct": f"{format_kpi_value(win_rate_pct, is_percentage=True)} ({total_winning_trades}/{total_trades})",
        "num_trades": total_trades,
        "applied_costs_info": f"매수/매도 수수료: {format_kpi_value(commission_rate_const*100, precision=3, is_percentage=True)}, 매도세: {format_kpi_value(sell_tax_rate_const*100, precision=3, is_percentage=True)}",
        "position_sizing_info": position_sizing_info_str, # 포트폴리오용 포지션 사이징 설명
        "conditions_applied": conditions_from_db # 적용된 공통 전략 조건
    }

    # 차트 데이터
    chart_labels_portfolio = [date.strftime('%Y-%m-%d') for date in portfolio_daily_values.index]
    chart_portfolio_values_portfolio = portfolio_daily_values['Portfolio_Value'].round(2).tolist()
    chart_drawdown_values_portfolio = portfolio_daily_values['Drawdown'].round(4).tolist()
    
    # 월별 수익률 히스토그램 (포트폴리오 전체 기준)
    monthly_portfolio_total_values = portfolio_daily_values['Portfolio_Value'].resample('M').last()
    monthly_returns_portfolio = monthly_portfolio_total_values.pct_change().fillna(0)
    histogram_frequencies_portfolio = []; histogram_labels_portfolio = []
    if not monthly_returns_portfolio.empty and len(monthly_returns_portfolio) > 1:
        hist, bin_edges = np.histogram(monthly_returns_portfolio * 100, bins=10)
        histogram_frequencies_portfolio = hist.tolist()
        for j in range(len(bin_edges) - 1): histogram_labels_portfolio.append(f"{bin_edges[j]:.1f}%~{bin_edges[j+1]:.1f}%")

    return {
        "kpis": kpis,
        "portfolio_df": portfolio_daily_values, # 전체 포트폴리오 가치, MDD 등
        "chart_labels": chart_labels_portfolio,
        "chart_portfolio_values": chart_portfolio_values_portfolio,
        "chart_drawdown_values": chart_drawdown_values_portfolio,
        "histogram_labels": histogram_labels_portfolio,
        "histogram_frequencies": histogram_frequencies_portfolio
        # 개별 자산별 결과도 필요시 반환 가능
    }
    
# --- (이전 코드: index, list_strategies 등은 거의 그대로 유지, stock_data_view도 유지) ---
@app.route('/')
def index():
    # 세션 등을 활용하여 로그인 상태에 따라 다른 페이지를 보여줄 수도 있습니다.
    return render_template('index.html') # 홈페이지 템플릿을 보여주도록 변경

@app.route('/strategies')
@login_required # 로그인한 사용자만 자신의 전략 목록을 볼 수 있도록
def list_strategies():
    conn = None
    cursor = None
    user_id = current_user.id # 현재 로그인한 사용자 ID
    try:
        conn = get_db_connection()
        if conn is None:
            flash("데이터베이스 연결에 실패했습니다.", "danger")
            return render_template('strategies_list.html', strategies=[])
        
        cursor = conn.cursor(dictionary=True)
        sql = """
            SELECT s.id, s.name, s.description, s.strategy_type, s.created_at, 
                   GROUP_CONCAT(CONCAT(c.indicator_type, ': ', c.value) SEPARATOR '; ') as conditions_summary
            FROM strategies s
            LEFT JOIN conditions c ON s.id = c.strategy_id
            WHERE s.user_id = %s  -- 현재 사용자의 전략만 가져오도록 수정
            GROUP BY s.id, s.name, s.description, s.strategy_type, s.created_at -- strategy_type 추가
            ORDER BY s.created_at DESC
        """
        cursor.execute(sql, (user_id,)) # user_id를 쿼리 파라미터로 전달
        strategies = cursor.fetchall()
        return render_template('strategies_list.html', strategies=strategies)
    except mysql.connector.Error as e:
        print(f"전략 목록 조회 오류: {e}")
        flash("전략 목록을 가져오는 중 오류가 발생했습니다.", "danger")
        return render_template('strategies_list.html', strategies=[])
    finally:
        if cursor:
            cursor.close()
        if conn and conn.is_connected():
            conn.close()


@app.route('/strategies/add', methods=['GET', 'POST'])
@login_required # 이 줄 추가
def add_strategy():
    if request.method == 'POST':
        strategy_name = request.form.get('name')
        strategy_description = request.form.get('description')
        strategy_type_from_form = request.form.get('strategy_type') # <<<<<< 이 줄 추가
        
        user_id = current_user.id

        # 동적으로 추가된 조건들 가져오기
        condition_indicator_types = []
        condition_values = []
        i = 0
        while True:
            # conditions-0-indicator_type, conditions-1-indicator_type ... 와 같은 형태로 들어옴
            indicator_type = request.form.get(f'conditions-{i}-indicator_type')
            value = request.form.get(f'conditions-{i}-value')
            if indicator_type is None and value is None: # 더 이상 조건이 없으면 중단
                # 혹은, 특정 필드(예: indicator_type) 하나만 체크해도 됨
                break
            if indicator_type: # indicator_type이 빈 문자열이 아닌 경우에만 유효한 조건으로 간주
                condition_indicator_types.append(indicator_type)
                condition_values.append(value if value is not None else '') # 값이 없으면 빈 문자열로
            i += 1
        
        if not strategy_name:
            flash("전략 이름은 필수입니다.", "warning")
            # 입력값 유지를 위해 템플릿에 데이터 다시 전달 (이 부분은 좀 더 정교하게 구현 필요)
            return render_template('add_strategy.html', name=strategy_name, description=strategy_description) 

        conn = None; cursor = None
        try:
            conn = get_db_connection()
            if conn is None: flash("DB 연결 실패", "danger"); return render_template('add_strategy.html', name=strategy_name, description=strategy_description)
            
            cursor = conn.cursor()
            # SQL INSERT 문에 strategy_type 및 user_id 추가
            sql_strategy = "INSERT INTO strategies (name, description, strategy_type, user_id) VALUES (%s, %s, %s, %s)" # <<<<<< 수정
            cursor.execute(sql_strategy, (strategy_name, strategy_description, strategy_type_from_form if strategy_type_from_form != "선택안함" else None, user_id)) # <<<<<< 수정
            strategy_id = cursor.lastrowid
            
            # 조건들을 conditions 테이블에 저장
            for idx in range(len(condition_indicator_types)):
                indicator_type = condition_indicator_types[idx]
                value = condition_values[idx]
                if indicator_type: # 유효한 지표 타입이 있을 경우에만 저장
                    sql_condition = "INSERT INTO conditions (strategy_id, indicator_type, value) VALUES (%s, %s, %s)"
                    cursor.execute(sql_condition, (strategy_id, indicator_type, value))
            
            conn.commit()
            flash(f"전략 '{strategy_name}'이(가) 성공적으로 등록되었습니다.", "success")
            return redirect(url_for('list_strategies'))
        except mysql.connector.Error as e:
            print(f"전략 등록 오류: {e}"); conn.rollback(); flash(f"등록 오류: {str(e)}", "danger")
            return render_template('add_strategy.html', name=strategy_name, description=strategy_description, strategy_type=strategy_type_from_form, strategy_types=STRATEGY_TYPES, user_id=user_id)
        finally:
            if cursor: cursor.close()
            if conn and conn.is_connected(): conn.close()
    else: # GET
        # GET 요청 시에도 strategy_types를 전달해야 <select>가 제대로 표시됨
        return render_template('add_strategy.html', strategy_types=STRATEGY_TYPES)
    
@app.route('/strategies/edit/<int:strategy_id>', methods=['GET', 'POST'])
def edit_strategy(strategy_id):
    conn = None; cursor = None
    user_id = current_user.id # 현재 로그인한 사용자 ID
    try:
        conn = get_db_connection()
        if conn is None: flash("DB 연결 실패", "danger"); return redirect(url_for('list_strategies'))
        cursor = conn.cursor(dictionary=True)

        if request.method == 'POST':
            new_name = request.form.get('name')
            new_description = request.form.get('description')
            new_strategy_type = request.form.get('strategy_type')

            # 동적으로 추가된 조건들 가져오기 (add_strategy와 동일한 로직)
            condition_indicator_types = []
            condition_values = []
            i = 0
            while True:
                indicator_type = request.form.get(f'conditions-{i}-indicator_type')
                value = request.form.get(f'conditions-{i}-value')
                if indicator_type is None and value is None: break
                if indicator_type:
                    condition_indicator_types.append(indicator_type)
                    condition_values.append(value if value is not None else '')
                i += 1

            if not new_name:
                flash("전략 이름은 필수입니다.", "warning")
                # 현재 정보 다시 로드하여 폼 보여주기
                cursor.execute("SELECT * FROM strategies WHERE id = %s", (strategy_id,))
                strategy = cursor.fetchone()
                cursor.execute("SELECT * FROM conditions WHERE strategy_id = %s ORDER BY id ASC", (strategy_id,))
                conditions_data = cursor.fetchall()
                if not strategy: flash("수정할 전략 없음", "danger"); return redirect(url_for('list_strategies'))
                return render_template('edit_strategy.html', strategy=strategy, conditions=conditions_data)

            # 1. strategies 테이블 업데이트 (strategy_type 추가)
            sql_update_strategy = "UPDATE strategies SET name = %s, description = %s, strategy_type = %s WHERE id = %s AND user_id = %s" # <<<<<< 수정
            cursor.execute(sql_update_strategy, (new_name, new_description, new_strategy_type if new_strategy_type != "선택안함" else None, strategy_id, user_id)) # <<<<<< 수정


            # 2. conditions 테이블 업데이트 (Delete-then-Insert 방식)
            sql_delete_old_conditions = "DELETE FROM conditions WHERE strategy_id = %s"
            cursor.execute(sql_delete_old_conditions, (strategy_id,))
            
            for idx in range(len(condition_indicator_types)):
                indicator_type = condition_indicator_types[idx]
                value = condition_values[idx]
                if indicator_type: # 유효한 지표 타입이 있을 경우에만 저장
                    sql_insert_condition = "INSERT INTO conditions (strategy_id, indicator_type, value) VALUES (%s, %s, %s)"
                    cursor.execute(sql_insert_condition, (strategy_id, indicator_type, value))
            
            conn.commit()
            flash(f"전략 '{new_name}'이(가) 성공적으로 수정되었습니다.", "success")
            return redirect(url_for('list_strategies'))

        else: # GET 요청
            cursor.execute("SELECT * FROM strategies WHERE id = %s", (strategy_id,))
            strategy = cursor.fetchone()
            if not strategy: flash("수정할 전략 없음", "danger"); return redirect(url_for('list_strategies'))
            
            cursor.execute("SELECT * FROM conditions WHERE strategy_id = %s ORDER BY id ASC", (strategy_id,))
            conditions_data = cursor.fetchall() # 모든 조건을 가져옴
            
            return render_template('edit_strategy.html', strategy=strategy, conditions=conditions_data, strategy_types=STRATEGY_TYPES)

    except mysql.connector.Error as e:
        print(f"전략 수정 중 DB 오류: {e}"); conn.rollback(); flash(f"DB 오류: {str(e)}", "danger")
    except Exception as e:
        print(f"전략 수정 중 일반 오류: {e}"); flash(f"일반 오류: {str(e)}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    return redirect(url_for('list_strategies')) # 오류 발생 시 목록으로


@app.route('/strategies/delete/<int:strategy_id>', methods=['POST'])
def delete_strategy(strategy_id):
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn is None:
            flash("데이터베이스 연결 실패", "danger")
            return redirect(url_for('list_strategies'))
        
        cursor = conn.cursor(dictionary=True)
        
        # (선택사항) 삭제 전 전략 이름 가져오기 (플래시 메시지용)
        cursor.execute("SELECT name FROM strategies WHERE id = %s", (strategy_id,))
        strategy = cursor.fetchone()
        strategy_name_for_flash = strategy['name'] if strategy else f"ID {strategy_id}"

        # strategies 테이블에서 해당 전략 삭제
        # ON DELETE CASCADE 제약 조건에 의해 conditions, results 테이블의 관련 레코드도 자동 삭제됨
        sql_delete_strategy = "DELETE FROM strategies WHERE id = %s"
        cursor.execute(sql_delete_strategy, (strategy_id,))
        
        # 삭제된 행의 수 확인 (선택적)
        if cursor.rowcount == 0:
            flash(f"ID {strategy_id}에 해당하는 전략을 찾을 수 없어 삭제하지 못했습니다.", "warning")
        else:
            conn.commit()
            flash(f"전략 '{strategy_name_for_flash}' (ID: {strategy_id})이(가) 성공적으로 삭제되었습니다.", "success")

        return redirect(url_for('list_strategies'))

    except mysql.connector.Error as e:
        print(f"전략 삭제 중 DB 오류: {e}")
        if conn: conn.rollback()
        flash(f"전략 삭제 중 오류 발생: {str(e)}", "danger")
        return redirect(url_for('list_strategies'))
    except Exception as e:
        print(f"전략 삭제 중 일반 오류: {e}")
        flash(f"전략 삭제 중 알 수 없는 오류 발생: {str(e)}", "danger")
        return redirect(url_for('list_strategies'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
        
# 백테스트 결과 목록 페이지 라우트 (새로 추가)
@app.route('/results')
def list_results():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn is None:
            flash("데이터베이스 연결에 실패하여 결과를 가져올 수 없습니다.", "danger")
            return render_template('results_list.html', results=[])

        cursor = conn.cursor(dictionary=True)
        # SELECT 쿼리에 r.cagr이 포함되어 있는지 확인!
        sql = """
            SELECT r.id, s.name as strategy_name, r.return_rate, r.mdd, r.win_rate, r.cagr, r.executed_at 
            FROM results r
            JOIN strategies s ON r.strategy_id = s.id
            ORDER BY r.executed_at DESC
        """
        cursor.execute(sql)
        results = cursor.fetchall() # 이 results 리스트의 각 항목이 res_item이 됩니다.
        return render_template('results_list.html', results=results)
    except mysql.connector.Error as e:
        print(f"백테스트 결과 목록 조회 오류: {e}")
        flash(f"백테스트 결과 목록을 가져오는 중 오류 발생: {str(e)}", "danger")
        return render_template('results_list.html', results=[])
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
        
# --- 주가 데이터 조회 기능 (stock_data_view)은 이전과 동일하게 유지 ---
@app.route('/stock_data', methods=['GET', 'POST'])
def stock_data_view():
    if request.method == 'POST':
        ticker = request.form.get('ticker')
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')

        if not ticker or not start_date_str or not end_date_str:
            return render_template('stock_data.html', error="종목코드와 기간을 모두 입력해주세요.", ticker=ticker, start_date=start_date_str, end_date=end_date_str)

        try:
            df = fdr.DataReader(ticker, start_date_str, end_date_str)
            if df.empty:
                return render_template('stock_data.html', error=f"{ticker}에 대한 데이터를 가져올 수 없습니다. 종목코드나 기간을 확인해주세요.", ticker=ticker, start_date=start_date_str, end_date=end_date_str)
            data_html = Markup(df.to_html(classes='table table-striped table-hover', border=0))
            return render_template('stock_data.html', data_table=data_html, ticker=ticker, start_date=start_date_str, end_date=end_date_str)
        except Exception as e:
            print(f"주가 데이터 조회 오류: {e}")
            return render_template('stock_data.html', error=f"데이터 조회 중 오류 발생: {e}", ticker=ticker, start_date=start_date_str, end_date=end_date_str)
            
    end_date_default = datetime.now().strftime('%Y-%m-%d')
    start_date_default = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    return render_template('stock_data.html', start_date=start_date_default, end_date=end_date_default)

@app.route('/backtest/run', methods=['GET'])
def run_backtest_page():
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn is None: flash("DB 연결 실패", "danger"); return redirect(url_for('home')) # 혹은 에러페이지
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name FROM strategies ORDER BY name ASC")
        strategies = cursor.fetchall()
        default_end_date = datetime.now().strftime('%Y-%m-%d')
        default_start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        return render_template('run_backtest.html', strategies=strategies, 
                               default_start_date=default_start_date, default_end_date=default_end_date)
    # ... (예외 처리) ...
    except Exception as e:
        flash(f"페이지 로드 중 오류: {str(e)}", "danger")
        return redirect(url_for('home'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()




# 백테스트 실행 로직 처리
# 백테스트 실행 로직 처리
@app.route('/backtest/execute', methods=['POST'])
@login_required # 사용자 계정 관리 기능과 연동 시
def execute_backtest():
    if request.method == 'POST':
        user_id = current_user.id # 사용자 ID 가져오기

        strategy_id = request.form.get('strategy_id')
        # 단일 티커 대신 여러 티커를 받도록 수정
        tickers_str = request.form.get('tickers') # '005930,035720,AAPL'
        weights_str = request.form.get('weights') # '60,40' 또는 비어있을 수 있음

        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        
        initial_capital_str = request.form.get('initial_capital', '10000000') # 초기 자본금

        position_sizing_method = request.form.get('position_sizing_method', 'all_in')
        fixed_amount_value_str = request.form.get('fixed_amount_value')
        fixed_percentage_value_str = request.form.get('fixed_percentage_value')

        if not all([strategy_id, tickers_str, start_date_str, end_date_str, initial_capital_str]):
            flash("전략, 티커 목록, 기간, 초기 자본금을 모두 올바르게 입력해주세요.", "warning")
            return redirect(url_for('run_backtest_page'))

        tickers = [t.strip().upper() for t in tickers_str.split(',') if t.strip()]
        if not tickers:
            flash("티커를 하나 이상 입력해주세요.", "warning")
            return redirect(url_for('run_backtest_page'))

        weights = []
        if weights_str:
            try:
                weights = [float(w.strip()) for w in weights_str.split(',') if w.strip()]
                if len(weights) != len(tickers):
                    flash("티커 수와 비중 수가 일치해야 합니다.", "warning")
                    return redirect(url_for('run_backtest_page'))
                if abs(sum(weights) - 100.0) > 0.01: # 부동소수점 오차 감안
                    flash(f"비중의 총합은 100%여야 합니다. (현재 총합: {sum(weights):.2f}%)", "warning")
                    return redirect(url_for('run_backtest_page'))
                weights = [w / 100.0 for w in weights] # 0.6, 0.4 형태로 변환
            except ValueError:
                flash("비중은 숫자로 입력해야 합니다.", "warning")
                return redirect(url_for('run_backtest_page'))
        else: # 비중 입력 없으면 동일 비중
            weights = [1.0 / len(tickers)] * len(tickers)

        try:
            initial_capital = float(initial_capital_str)
            if initial_capital <= 0:
                flash("초기 자본금은 0보다 커야 합니다.", "danger")
                return redirect(url_for('run_backtest_page'))

            fixed_amount_param = float(fixed_amount_value_str) if fixed_amount_value_str else 1000000.0
            # 포트폴리오에서는 고정 비율 투자가 전체 포트폴리오 자산 대비인지,
            # 아니면 개별 종목에 할당된 자산 대비인지 명확히 해야 합니다.
            # 여기서는 'all_in' (자산별 할당된 금액 내에서 전량) 또는 'fixed_amount' (자산별 고정 금액)을 우선 고려합니다.
            # 'fixed_percentage'는 전체 포트폴리오 자산 대비로 간주하고, 각 자산에 분배하는 로직이 추가로 필요합니다. (복잡도 증가)
            # 단순화를 위해, 포지션 사이징은 각 자산별로 할당된 초기 자본 내에서 이루어진다고 가정합니다.
            fixed_percentage_param = float(fixed_percentage_value_str) / 100.0 if fixed_percentage_value_str else 0.1 

        except ValueError:
            flash("숫자 파라미터(초기 자본금, 포지션 사이징 값)가 잘못되었습니다.", "danger")
            return redirect(url_for('run_backtest_page'))
        
        # ... (DB에서 전략 정보, 조건 정보 가져오는 부분은 기존과 유사) ...
        conn = None
        cursor = None
        try:
            conn = get_db_connection()
            if conn is None:
                flash("DB 연결 실패", "danger")
                return "DB 연결 실패", 500
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT name FROM strategies WHERE id = %s AND (user_id = %s OR user_id IS NULL)", (strategy_id, user_id)) # 사용자 데이터 접근 제어
            strategy_info = cursor.fetchone()
            if not strategy_info:
                flash("선택한 전략을 찾을 수 없습니다.", "danger")
                return redirect(url_for('run_backtest_page'))
            
            cursor.execute("SELECT id, indicator_type, value FROM conditions WHERE strategy_id = %s", (strategy_id,))
            conditions_from_db = cursor.fetchall()
            if not conditions_from_db:
                flash(f"전략 '{strategy_info['name']}'에 설정된 조건이 없습니다.", "warning")
                return redirect(url_for('run_backtest_page'))

            # --- 포트폴리오 백테스트 로직 호출 ---
            # portfolio_backtest_logic 함수는 새로 만들어야 합니다.
            portfolio_results = portfolio_backtest_logic(
                tickers, weights, start_date_str, end_date_str,
                conditions_from_db, initial_capital,
                position_sizing_method, fixed_amount_param, fixed_percentage_param, # 이 부분은 포트폴리오 맥락에 맞게 재해석 필요
                COMMISSION_RATE, SELL_TAX_RATE, strategy_info['name']
            )

            if portfolio_results is None: # 오류 발생 시
                # flash 메시지는 portfolio_backtest_logic 내부 또는 여기서 처리
                return redirect(url_for('run_backtest_page'))

            # portfolio_results에서 kpis, portfolio_df, chart_data 등을 받아옵니다.
            kpis = portfolio_results['kpis']
            portfolio_df = portfolio_results['portfolio_df'] # 포트폴리오 전체 가치 변화 등이 담긴 DataFrame
            chart_labels = portfolio_results['chart_labels']
            chart_portfolio_values = portfolio_results['chart_portfolio_values']
            chart_drawdown_values = portfolio_results['chart_drawdown_values']
            # 개별 종목 데이터 및 히스토그램은 복잡해지므로 우선 포트폴리오 전체 지표에 집중

            # DB 저장용 파라미터 구성
            backtest_params_for_db = {
                "strategy_id": int(strategy_id),
                "strategy_name": strategy_info['name'],
                "tickers": tickers, # 리스트 형태
                "weights": [w * 100 for w in weights], # % 형태로 저장
                "start_date": start_date_str,
                "end_date": end_date_str,
                "conditions": conditions_from_db,
                "initial_capital": initial_capital,
                "position_sizing_method": position_sizing_method,
                "fixed_amount_value": fixed_amount_param if position_sizing_method == 'fixed_amount' else None,
                "fixed_percentage_value": fixed_percentage_param * 100 if position_sizing_method == 'fixed_percentage' else None,
                "commission_rate": COMMISSION_RATE,
                "sell_tax_rate": SELL_TAX_RATE,
                "is_portfolio": True # 포트폴리오 백테스트임을 명시
            }
            
            # DB 저장 로직 (기존과 유사하나, results 테이블에 포트폴리오 관련 정보 저장 방식 고려 필요)
            # 예: ticker 컬럼 대신 tickers (JSON 문자열), weights (JSON 문자열) 컬럼 추가 등
            # 또는 parameters JSON에 모두 저장
            executed_at_dt = datetime.now()
            # cagr_to_db 등 KPI 값 추출 (기존 로직 참고)
            cagr_to_db = float(kpis["cagr_pct"].replace('%','')) if isinstance(kpis["cagr_pct"], str) and '%' in kpis["cagr_pct"] and "N/A" not in kpis["cagr_pct"] else None
            # ... (다른 kpi 값들도 숫자형으로 변환) ...

            sql_insert_result = """
                INSERT INTO results (strategy_id, user_id, return_rate, mdd, win_rate, cagr, executed_at, parameters) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql_insert_result, (
                int(strategy_id), 
                user_id,
                float(kpis["total_return_pct"].replace('%','')), 
                float(kpis["max_drawdown_pct"].replace('%','')), 
                float(kpis["win_rate_pct"].split('%')[0]) if kpis["win_rate_pct"] else 0.0,
                cagr_to_db,
                executed_at_dt,
                json.dumps(backtest_params_for_db, ensure_ascii=False, default=str) 
            ))
            conn.commit()
            new_result_id = cursor.lastrowid

            # 결과 표시 (backtest_result.html 템플릿은 포트폴리오 결과에 맞게 일부 수정 필요)
            # 예를 들어, 종목별 상세 데이터 대신 포트폴리오 전체 통계 위주로
            return render_template('backtest_result.html', 
                                   result=kpis, # 포트폴리오 전체 KPI
                                   result_id=new_result_id,
                                   # 포트폴리오 전체 가치 변화, 혹은 주요 자산 데이터 요약
                                   stock_data_html=Markup(portfolio_df[['Portfolio_Value', 'Drawdown']].tail(20).to_html(classes="table table-sm table-striped table-hover", float_format='{:,.2f}'.format)),
                                   chart_labels=chart_labels, 
                                   chart_portfolio_values=chart_portfolio_values,
                                   chart_drawdown_values=chart_drawdown_values,
                                   # 포트폴리오의 월별 수익률 히스토그램 (portfolio_results에서 받아와야 함)
                                   histogram_labels=portfolio_results.get('histogram_labels', []), 
                                   histogram_frequencies=portfolio_results.get('histogram_frequencies', []),
                                   is_portfolio_result=True # 템플릿에서 구분용
                                  )

        # ... (기존 예외 처리 부분) ...
        except ValueError as ve:
            flash(f"백테스트 설정 오류: {str(ve)}", "danger")
            return redirect(url_for('run_backtest_page'))
        except mysql.connector.Error as e:
            print(f"DB 오류: {e}"); flash(f"DB 오류: {e}", "danger")
            if conn: conn.rollback()
            return redirect(url_for('run_backtest_page'))
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"일반 오류: {type(e).__name__}: {e}"); flash(f"일반 오류: {type(e).__name__} - {e}", "danger")
            return redirect(url_for('run_backtest_page'))
        finally:
            if cursor: cursor.close()
            if conn and conn.is_connected(): conn.close()
            
    return redirect(url_for('run_backtest_page'))

# SP_Project/app.py

@app.route('/results/view/<int:result_id>')
@login_required # 사용자 인증 추가
def view_result(result_id):
    user_id = current_user.id # 현재 사용자 ID
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn is None:
            flash("DB 연결 실패", "danger")
            return redirect(url_for('list_results'))
        cursor = conn.cursor(dictionary=True)
        
        # 현재 사용자의 결과만 조회하도록 수정
        cursor.execute("SELECT * FROM results WHERE id = %s AND user_id = %s", (result_id, user_id))
        saved_result = cursor.fetchone()
        
        if not saved_result:
            flash("해당 ID의 백테스트 결과를 찾을 수 없거나 접근 권한이 없습니다.", "danger")
            return redirect(url_for('list_results'))

        if not saved_result['parameters']:
            flash("저장된 실행 파라미터가 없어 상세 결과를 재현할 수 없습니다. (기본 정보만 표시)", "warning")
            # parameters가 없는 경우, DB에 저장된 기본 KPI만 표시 (이전 fallback 로직과 유사하게)
            kpis_fallback = {
                "strategy_name": "N/A", # 전략 이름은 strategies 테이블에서 가져와야 함
                "ticker_display": "N/A",
                "period": "N/A",
                "conditions_applied": [],
                "initial_capital": format_kpi_value(saved_result.get('initial_capital'), is_currency=True, default_na_str="정보 없음"), # DB에 initial_capital이 없다면 추가 필요
                "final_portfolio_value": format_kpi_value(saved_result.get('final_value'), is_currency=True, default_na_str="정보 없음"), # DB에 final_value가 없다면 추가 필요
                "total_return_pct": format_kpi_value(saved_result.get('return_rate'), is_percentage=True),
                "max_drawdown_pct": format_kpi_value(saved_result.get('mdd'), is_percentage=True),
                "win_rate_pct": format_kpi_value(saved_result.get('win_rate'), is_percentage=True) + f" ({saved_result.get('winning_trades',0)}/{saved_result.get('total_trades',0)})", # trade 횟수 DB 저장 고려
                "cagr_pct": format_kpi_value(saved_result.get('cagr'), is_percentage=True, default_na_str="N/A"),
                "num_trades": saved_result.get('total_trades',0),
                "position_sizing_info": "정보 없음",
                "applied_costs_info": "정보 없음"
            }
            # 전략 이름 가져오기
            cursor.execute("SELECT name, strategy_type FROM strategies WHERE id = %s", (saved_result['strategy_id'],))
            strategy_info_fallback = cursor.fetchone()
            if strategy_info_fallback:
                kpis_fallback["strategy_name"] = strategy_info_fallback['name']
                kpis_fallback["strategy_type"] = strategy_info_fallback.get('strategy_type')

            return render_template('backtest_result.html', result=kpis_fallback, result_id=result_id,
                                   stock_data_html="상세 데이터 재구성 불가",
                                   chart_labels=[], chart_portfolio_values=[], chart_drawdown_values=[],
                                   histogram_labels=[], histogram_frequencies=[],
                                   is_portfolio_result=False) # 또는 params가 없으면 is_portfolio도 알 수 없음

        params = json.loads(saved_result['parameters'])
        is_portfolio = params.get("is_portfolio", False) # 포트폴리오 여부 확인

        kpis = {}
        stock_data_df_processed = pd.DataFrame()
        display_cols = []
        chart_labels = []
        chart_portfolio_values = []
        chart_drawdown_values = []
        histogram_labels = []
        histogram_frequencies = []

        if is_portfolio:
            # 포트폴리오 결과 재구성
            # portfolio_backtest_logic 함수가 kpis, df, 차트 데이터 등을 포함한 dict를 반환한다고 가정
            portfolio_results = portfolio_backtest_logic(
                params['tickers'], # 'tickers' (복수형) 사용
                [w/100.0 for w in params['weights']], # 저장된 weights가 %라면 0-1 범위로 변환
                params['start_date'], params['end_date'], params['conditions'],
                params['initial_capital'], params['position_sizing_method'],
                params.get('fixed_amount_value'), params.get('fixed_percentage_value'),
                params['commission_rate'], params['sell_tax_rate'],
                params['strategy_name'] # 전략 이름 전달
            )
            if not portfolio_results:
                flash("포트폴리오 결과 상세 보기 중 오류 발생.", "danger")
                return redirect(url_for('list_results'))

            kpis = portfolio_results['kpis']
            stock_data_df_processed = portfolio_results['portfolio_df']
            chart_labels = portfolio_results['chart_labels']
            chart_portfolio_values = portfolio_results['chart_portfolio_values']
            chart_drawdown_values = portfolio_results['chart_drawdown_values']
            histogram_labels = portfolio_results.get('histogram_labels', [])
            histogram_frequencies = portfolio_results.get('histogram_frequencies', [])
            
            # 포트폴리오의 경우 display_cols 설정
            if not stock_data_df_processed.empty:
                 display_cols = [col for col in ['Portfolio_Value', 'Peak', 'Drawdown'] if col in stock_data_df_processed.columns]
            # kpis에 티커 표시용 문자열 추가 (portfolio_backtest_logic에서 이미 처리했을 수 있음)
            if 'ticker_display' not in kpis:
                 kpis['ticker_display'] = ", ".join(params['tickers'])
            
        else: # 단일 종목 결과 재구성
            # 단일 종목의 경우 params에 'ticker'가 있어야 함
            if 'ticker' not in params:
                flash("단일 종목 백테스트 파라미터에 'ticker' 정보가 없습니다. (저장된 데이터 오류)", "danger")
                return redirect(url_for('list_results'))

            kpis_single, stock_data_df_processed_single, display_cols_single, chart_labels_single, chart_portfolio_values_single, chart_drawdown_values_single, hist_lbls_single, hist_freqs_single = \
                run_backtest_logic(
                    params['ticker'], # 'ticker' (단수형) 사용
                    params['start_date'], params['end_date'], params['conditions'],
                    params['initial_capital'], params['position_sizing_method'],
                    params.get('fixed_amount_value'), params.get('fixed_percentage_value'),
                    params['commission_rate'], params['sell_tax_rate']
                )
            kpis = kpis_single
            stock_data_df_processed = stock_data_df_processed_single
            display_cols = display_cols_single
            chart_labels = chart_labels_single
            chart_portfolio_values = chart_portfolio_values_single
            chart_drawdown_values = chart_drawdown_values_single
            histogram_labels = hist_lbls_single
            histogram_frequencies = hist_freqs_single

            # kpis 딕셔너리에 추가 정보 병합
            kpis["strategy_name"] = params['strategy_name']
            kpis["ticker_display"] = params['ticker'] # 단일 티커 표시용
            kpis["period"] = f"{params['start_date']} ~ {params['end_date']}"
            kpis["conditions_applied"] = params['conditions']
            # (선택) 전략 유형 정보도 kpis에 추가 (params 또는 DB에서 가져오기)
            kpis["strategy_type"] = params.get('strategy_type')
            if not kpis["strategy_type"]: # params에 없다면 DB에서 다시 조회
                cursor.execute("SELECT strategy_type FROM strategies WHERE id = %s", (saved_result['strategy_id'],))
                strat_info = cursor.fetchone()
                if strat_info: kpis["strategy_type"] = strat_info.get('strategy_type')


        # stock_data_html 생성 (display_cols이 비어있지 않은 경우에만)
        stock_data_html_content = "표시할 데이터가 없습니다."
        if display_cols and not stock_data_df_processed.empty:
            valid_display_cols = [col for col in display_cols if col in stock_data_df_processed.columns]
            if valid_display_cols :
                stock_data_html_content = Markup(stock_data_df_processed[valid_display_cols].tail(20).to_html(classes="table table-sm table-striped table-hover", float_format='{:,.2f}'.format))
        elif not stock_data_df_processed.empty and 'Portfolio_Value' in stock_data_df_processed.columns: # display_cols이 없어도 기본 컬럼 표시
             stock_data_html_content = Markup(stock_data_df_processed[['Portfolio_Value']].tail(20).to_html(classes="table table-sm table-striped table-hover", float_format='{:,.2f}'.format))


        return render_template('backtest_result.html', 
                               result=kpis, 
                               result_id=result_id,
                               stock_data_html=stock_data_html_content,
                               chart_labels=chart_labels, 
                               chart_portfolio_values=chart_portfolio_values,
                               chart_drawdown_values=chart_drawdown_values,
                               histogram_labels=histogram_labels, 
                               histogram_frequencies=histogram_frequencies,
                               is_portfolio_result=is_portfolio # 템플릿에 전달
                              )

    except KeyError as ke: # 특정 키 에러 처리
        flash(f"결과 상세 보기 중 필요한 데이터(키: {ke})를 찾을 수 없습니다. 저장된 파라미터를 확인해주세요.", "danger")
        print(f"KeyError in view_result: {ke}")
        import traceback; traceback.print_exc()
        return redirect(url_for('list_results'))
    except ValueError as ve:
        flash(f"결과 조회 중 데이터 변환 오류: {str(ve)}", "danger")
        return redirect(url_for('list_results'))
    except mysql.connector.Error as e:
        print(f"DB 오류 (view_result): {e}"); flash(f"DB 오류: {e}", "danger")
        return redirect(url_for('list_results'))
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"일반 오류 (view_result): {type(e).__name__}: {e}"); flash(f"일반 오류: {type(e).__name__} - {e}", "danger")
        return redirect(url_for('list_results'))
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

@app.route('/results/download/excel/<int:result_id>')
@login_required # 사용자 인증 추가
def download_excel_report(result_id):
    user_id = current_user.id # 현재 사용자 ID
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn is None:
            flash("DB 연결 실패로 리포트를 생성할 수 없습니다.", "danger")
            return redirect(url_for('list_results')) # view_result 대신 list_results로 리다이렉트할 수도 있음
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM results WHERE id = %s AND user_id = %s", (result_id, user_id)) # 사용자 ID로 필터링
        saved_result = cursor.fetchone()

        if not saved_result or not saved_result['parameters']:
            flash("리포트를 생성할 상세 파라미터 정보가 없습니다.", "warning")
            # result_id가 유효하다면 view_result로, 아니면 list_results로
            return redirect(url_for('view_result', result_id=result_id) if result_id else url_for('list_results'))

        params = json.loads(saved_result['parameters'])
        
        is_portfolio = params.get("is_portfolio", False) # 포트폴리오 여부 확인

        kpis = {}
        stock_data_df_processed = pd.DataFrame() # 초기화
        display_cols = [] # 초기화

        if is_portfolio:
            # 포트폴리오 결과 재구성
            portfolio_results = portfolio_backtest_logic(
                params['tickers'], # 'tickers' (복수형) 사용
                [w/100.0 for w in params['weights']],
                params['start_date'], params['end_date'], params['conditions'],
                params['initial_capital'], params['position_sizing_method'],
                params.get('fixed_amount_value'), params.get('fixed_percentage_value'),
                params['commission_rate'], params['sell_tax_rate'],
                params['strategy_name']
            )
            if not portfolio_results:
                flash("포트폴리오 결과 재구성 중 오류 발생.", "danger")
                return redirect(url_for('view_result', result_id=result_id))
            
            kpis = portfolio_results['kpis']
            stock_data_df_processed = portfolio_results['portfolio_df'] # 포트폴리오 전체 데이터
            # 포트폴리오의 경우 display_cols을 적절히 설정 (예: 포트폴리오 가치, 드로우다운 등)
            if not stock_data_df_processed.empty:
                 display_cols = [col for col in ['Portfolio_Value', 'Peak', 'Drawdown', 'Signal'] if col in stock_data_df_processed.columns]
            # kpis 딕셔너리에 'ticker_display' 추가 (포트폴리오 로직에서 이미 처리되었을 수 있음)
            if 'ticker' not in kpis: # portfolio_backtest_logic에서 kpis['ticker']로 티커목록을 넣었다면
                kpis['ticker_display'] = ", ".join(params['tickers'])


        else: # 단일 종목 결과 재구성
            # 단일 종목의 경우 params에 'ticker'가 있어야 함
            if 'ticker' not in params:
                flash("단일 종목 백테스트 파라미터에 'ticker' 정보가 없습니다.", "danger")
                return redirect(url_for('view_result', result_id=result_id))

            kpis_single, stock_data_df_processed_single, display_cols_single, _, _, _, _, _ = \
                run_backtest_logic(
                    params['ticker'], # 'ticker' (단수형) 사용
                    params['start_date'], params['end_date'], params['conditions'],
                    params['initial_capital'], params['position_sizing_method'],
                    params.get('fixed_amount_value'), params.get('fixed_percentage_value'),
                    params['commission_rate'], params['sell_tax_rate']
                )
            kpis = kpis_single # 단일 kpi 할당
            stock_data_df_processed = stock_data_df_processed_single
            display_cols = display_cols_single
            
            # kpis 딕셔너리에 추가 정보 병합
            kpis["strategy_name"] = params['strategy_name']
            kpis["ticker_display"] = params['ticker'] # 단일 티커 표시용
            kpis["period"] = f"{params['start_date']} ~ {params['end_date']}"
            kpis["conditions_applied_display"] = ", ".join([f"{c['indicator_type']}:{c['value']}" for c in params['conditions']])


        # --- Excel 파일 생성 시작 ---
        wb = Workbook()
        ws = wb.active
        ws.title = "백테스트 요약"

        # 스타일 정의 (기존과 동일)
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        # border_bottom_thin = Border(bottom=Side(style='thin')) # 사용되지 않으면 제거 가능

        # A. 보고서 제목 및 기본 정보
        # 파일명 생성 시 사용할 티커 정보 (포트폴리오/단일 구분)
        ticker_info_for_title = kpis.get('ticker_display', params.get('ticker', 'N/A') if not is_portfolio else ", ".join(params.get('tickers', ['N/A'])))
        
        ws['A1'] = f"{params['strategy_name']} ({ticker_info_for_title}) 백테스트 결과"
        ws.merge_cells('A1:D1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        ws['A2'] = f"기간: {params['start_date']} ~ {params['end_date']}"
        ws.merge_cells('A2:D2')
        
        # 조건 표시 (kpis에서 가져오거나 params에서 재구성)
        conditions_display = kpis.get('conditions_applied_display', ", ".join([f"{c['indicator_type']}:{c['value']}" for c in params.get('conditions', [])]))
        ws['A3'] = f"조건: {conditions_display}"
        ws.merge_cells('A3:D3')
        
        ws['A4'] = f"포지션 사이징: {kpis.get('position_sizing_info', '정보 없음')}" # kpis에 있어야 함
        ws.merge_cells('A4:D4')
        ws['A5'] = f"적용 비용: {kpis.get('applied_costs_info', '정보 없음')}" # kpis에 있어야 함
        ws.merge_cells('A5:D5')
        
        current_row = 7

        # B. 주요 성과 지표 (KPI) - kpis 딕셔너리 사용
        ws[f'A{current_row}'] = "주요 성과 지표"
        ws[f'A{current_row}'].font = subheader_font
        current_row += 1
        
        kpi_data_source = [
            ("초기 자본금", kpis.get('initial_capital', 'N/A')),
            ("최종 자산 가치", kpis.get('final_portfolio_value', 'N/A')),
            ("누적 수익률 (%)", kpis.get('total_return_pct', '0.00%').replace('%','')),
            ("CAGR (%)", kpis.get('cagr_pct', 'N/A').replace('%','') if isinstance(kpis.get('cagr_pct'), str) and "N/A" not in kpis.get('cagr_pct', '') else kpis.get('cagr_pct', 'N/A')),
            ("최대 낙폭 (MDD) (%)", kpis.get('max_drawdown_pct', '0.00%').replace('%','')),
            ("승률 (%)", kpis.get('win_rate_pct', '0.00% (0/0)').split('%')[0] if '%' in kpis.get('win_rate_pct', '') else kpis.get('win_rate_pct', '0 (0/0)').split(' ')[0]),
            ("총 거래 횟수", kpis.get('num_trades', 0))
        ]
        for desc, value_str in kpi_data_source:
            ws[f'A{current_row}'] = desc
            # 숫자 변환 시도 (오류 방지 강화)
            try:
                if isinstance(value_str, str) and ("원" in value_str or "," in value_str): # 통화 형식 처리
                     value_str_cleaned = value_str.replace(' 원','').replace(',','')
                     value = float(value_str_cleaned) if value_str_cleaned.replace('.', '', 1).replace('-', '', 1).isdigit() else value_str
                elif isinstance(value_str, str) and value_str.replace('.', '', 1).replace('-', '', 1).isdigit():
                    value = float(value_str)
                elif isinstance(value_str, (int, float)):
                    value = value_str
                else: # 변환 어려운 경우 문자열 그대로
                    value = value_str
            except ValueError:
                value = value_str # 변환 실패 시 원본 문자열

            ws[f'B{current_row}'] = value
            if isinstance(ws[f'B{current_row}'].value, (int, float)):
                 ws[f'B{current_row}'].number_format = '#,##0.00' if (isinstance(value, float) and (abs(value - int(value)) > 0.0001 or "." in str(value_str))) else '#,##0'
            current_row += 1
        current_row += 1

        # C. 일별 데이터 시트 추가
        if not stock_data_df_processed.empty and display_cols:
            ws_daily = wb.create_sheet(title="일별 데이터")
            
            # display_cols에 있는 컬럼만 추출하되, 없는 컬럼은 무시
            valid_display_cols = [col for col in display_cols if col in stock_data_df_processed.columns]
            if not valid_display_cols and 'Close' in stock_data_df_processed.columns : # display_cols이 비었지만 Close라도 있으면
                valid_display_cols = ['Close', 'Portfolio_Value', 'Drawdown'] # 기본값
                valid_display_cols = [col for col in valid_display_cols if col in stock_data_df_processed.columns]


            if valid_display_cols:
                daily_data_to_export = stock_data_df_processed[valid_display_cols]
                
                for r_idx, row in enumerate(dataframe_to_rows(daily_data_to_export, index=True, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws_daily.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1: # Header
                            cell.font = header_font
                            cell.alignment = center_align
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00' # 소수점 두 자리로 통일 또는 조건부 포맷팅

                # 컬럼 너비 자동 조절 (기존과 동일)
                for col_obj in ws_daily.columns: # col_obj 사용 (이전 코드에서 col이었음)
                    max_length = 0
                    column_letter = col_obj[0].column_letter # Get the column letter
                    for cell in col_obj:
                        try:
                            if cell.value is not None and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) if max_length > 0 else 12 # 최소 너비
                    ws_daily.column_dimensions[column_letter].width = adjusted_width
            else:
                ws_daily['A1'] = "표시할 일별 데이터 컬럼이 없습니다."


        # --- 파일 저장 및 전송 ---
        virtual_workbook_stream = io.BytesIO()
        wb.save(virtual_workbook_stream)
        virtual_workbook_stream.seek(0)

        # 파일명 생성 (포트폴리오/단일 구분)
        filename_ticker_part = params.get('ticker', 'portfolio') if not is_portfolio else "_".join(params.get('tickers', ['portfolio']))
        # 파일명 길이 및 특수문자 처리 강화
        safe_strategy_name = "".join(c if c.isalnum() else '_' for c in params['strategy_name'])[:30]
        safe_ticker_part = "".join(c if c.isalnum() else '_' for c in filename_ticker_part)[:30]

        report_filename = f"backtest_report_{safe_strategy_name}_{safe_ticker_part}_{result_id}.xlsx"
        
        return send_file(
            virtual_workbook_stream,
            as_attachment=True,
            download_name=report_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except KeyError as ke: # 특정 키 에러 처리
        flash(f"리포트 생성 중 필요한 데이터(키: {ke})가 부족합니다. 저장된 파라미터를 확인해주세요.", "danger")
        print(f"KeyError in download_excel_report: {ke}")
        import traceback; traceback.print_exc()
    except ValueError as ve:
        flash(f"리포트 생성 중 데이터 변환 오류: {str(ve)}", "danger")
        print(f"ValueError in download_excel_report: {ve}")
    except mysql.connector.Error as e:
        print(f"DB 오류 (download_excel): {e}"); flash(f"DB 오류: {e}", "danger")
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"일반 오류 (download_excel): {type(e).__name__}: {e}"); flash(f"일반 오류: {type(e).__name__} - {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    
    # 오류 발생 시 리다이렉트
    return redirect(url_for('view_result', result_id=result_id) if result_id else url_for('list_results'))

# --- (home, index 등 다른 라우트들은 이전과 동일) ---
@app.route('/home')
def home():
    return render_template('index.html')

@app.route('/') # 루트 경로를 home으로 연결
def root_redirect():
    return redirect(url_for('home'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        flash("이미 로그인되어 있습니다.", "info")
        return redirect(url_for('home'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if not username or not password or not confirm_password:
            flash("모든 필드를 입력해주세요.", "danger")
            return render_template('register.html')

        if password != confirm_password:
            flash("비밀번호가 일치하지 않습니다.", "danger")
            return render_template('register.html', username=username)

        conn = None
        cursor = None
        try:
            conn = get_db_connection()
            if conn is None:
                flash("데이터베이스 연결에 실패했습니다.", "danger")
                return render_template('register.html', username=username)
            
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
            existing_user = cursor.fetchone()

            if existing_user:
                flash("이미 사용 중인 사용자명입니다.", "warning")
                return render_template('register.html', username=username)

            hashed_password = generate_password_hash(password)
            cursor.execute("INSERT INTO users (username, password) VALUES (%s, %s)", (username, hashed_password))
            conn.commit()
            
            flash(f"{username}님, 회원가입이 완료되었습니다. 로그인해주세요.", "success")
            return redirect(url_for('login'))
        except mysql.connector.Error as e:
            print(f"회원가입 중 DB 오류: {e}")
            if conn: conn.rollback()
            flash("회원가입 중 오류가 발생했습니다. 다시 시도해주세요.", "danger")
            return render_template('register.html', username=username)
        finally:
            if cursor: cursor.close()
            if conn and conn.is_connected(): conn.close()
            
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        flash("이미 로그인되어 있습니다.", "info")
        return redirect(url_for('home'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            flash("사용자명과 비밀번호를 모두 입력해주세요.", "danger")
            return render_template('login.html', username=username)

        conn = None
        cursor = None
        try:
            conn = get_db_connection()
            if conn is None:
                flash("데이터베이스 연결에 실패했습니다.", "danger")
                return render_template('login.html', username=username)

            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
            user_data = cursor.fetchone()

            if user_data and check_password_hash(user_data['password'], password):
                user_obj = User(id=user_data['id'], username=user_data['username'])
                login_user(user_obj) # Flask-Login을 통해 사용자 로그인 처리
                flash(f"{user_data['username']}님, 환영합니다!", "success")
                
                # 로그인 후 이동할 다음 페이지가 있다면 그곳으로, 없다면 홈으로
                next_page = request.args.get('next')
                return redirect(next_page or url_for('home'))
            else:
                flash("사용자명이 존재하지 않거나 비밀번호가 틀렸습니다.", "danger")
                return render_template('login.html', username=username)
        except mysql.connector.Error as e:
            print(f"로그인 중 DB 오류: {e}")
            flash("로그인 중 오류가 발생했습니다. 다시 시도해주세요.", "danger")
            return render_template('login.html', username=username)
        finally:
            if cursor: cursor.close()
            if conn and conn.is_connected(): conn.close()

    return render_template('login.html')

@app.route('/results/download/pdf/<int:result_id>')
@login_required
def download_pdf_report(result_id):
    user_id = current_user.id
    conn = None; cursor = None
    try:
        conn = get_db_connection()
        if conn is None:
            flash("DB 연결 실패로 PDF 리포트를 생성할 수 없습니다.", "danger")
            return redirect(url_for('view_result', result_id=result_id) or url_for('list_results'))
        cursor = conn.cursor(dictionary=True)

        # 1. 결과 데이터 및 파라미터 가져오기 (Excel 다운로드와 유사)
        cursor.execute("SELECT * FROM results WHERE id = %s AND user_id = %s", (result_id, user_id))
        saved_result = cursor.fetchone()

        if not saved_result or not saved_result['parameters']:
            flash("PDF 리포트를 생성할 상세 파라미터 정보가 없습니다.", "warning")
            return redirect(url_for('view_result', result_id=result_id) or url_for('list_results'))

        params = json.loads(saved_result['parameters'])

        # 2. 파라미터를 사용하여 백테스트 로직 재실행 또는 데이터 재구성
        #    run_backtest_logic 또는 portfolio_backtest_logic 호출
        #    Excel 생성 로직과 매우 유사하게 kpis, stock_data_df_processed 등을 얻습니다.
        
        is_portfolio = params.get("is_portfolio", False)
        equity_chart_base64 = None
        drawdown_chart_base64 = None
        
        if is_portfolio:
            # 포트폴리오 로직 호출 (portfolio_backtest_logic이 kpis, df, 차트 데이터 반환 가정)
            portfolio_results = portfolio_backtest_logic(
                params['tickers'], [w/100.0 for w in params['weights']], # weights는 %로 저장되어 있을 수 있으므로 0-1 범위로 변환
                params['start_date'], params['end_date'], params['conditions'],
                params['initial_capital'], params['position_sizing_method'],
                params.get('fixed_amount_value'), params.get('fixed_percentage_value'),
                params['commission_rate'], params['sell_tax_rate'],
                params['strategy_name']
            )
            if not portfolio_results:
                flash("포트폴리오 결과 재구성 중 오류 발생.", "danger")
                return redirect(url_for('view_result', result_id=result_id))

            kpis_for_report = portfolio_results['kpis']
            # stock_data_df_for_report = portfolio_results['portfolio_df'] # 포트폴리오 전체 데이터
            chart_labels = portfolio_results['chart_labels']
            chart_portfolio_values = portfolio_results['chart_portfolio_values']
            chart_drawdown_values = portfolio_results['chart_drawdown_values']
            
            kpis_for_report['ticker_display'] = ", ".join(params['tickers'])

        else: # 단일 종목
            kpis_for_report, stock_data_df_processed, display_cols, chart_labels, chart_portfolio_values, chart_drawdown_values, _, _ = \
                run_backtest_logic(
                    params['ticker'], params['start_date'], params['end_date'], params['conditions'],
                    params['initial_capital'], params['position_sizing_method'],
                    params.get('fixed_amount_value'), params.get('fixed_percentage_value'),
                    params['commission_rate'], params['sell_tax_rate']
                )
            # kpis에 추가 정보 병합
            kpis_for_report["strategy_name"] = params['strategy_name']
            kpis_for_report["ticker_display"] = params['ticker'] # 단일 티커
            kpis_for_report["period"] = f"{params['start_date']} ~ {params['end_date']}"
            kpis_for_report["conditions_applied"] = params['conditions']
            # stock_data_df_for_report = stock_data_df_processed[display_cols]

        # KPI 값 중 % 제거 및 숫자형으로 변환 (템플릿에서 사용하기 위함)
        kpis_for_report['total_return_pct_raw'] = float(kpis_for_report['total_return_pct'].replace('%','')) if '%' in kpis_for_report['total_return_pct'] else 0.0
        kpis_for_report['cagr_pct_raw'] = float(kpis_for_report['cagr_pct'].replace('%','')) if '%' in kpis_for_report['cagr_pct'] and "N/A" not in kpis_for_report['cagr_pct'] else 0.0
        kpis_for_report['strategy_type'] = params.get('strategy_type', saved_result.get('strategy_type', None)) # 전략 유형 정보 추가

        # 3. 차트 이미지 생성 (base64)
        if chart_labels and chart_portfolio_values:
            equity_chart_base64 = generate_chart_for_pdf(chart_labels, chart_portfolio_values, "포트폴리오 가치 변화", "가치", "equity")
        if chart_labels and chart_drawdown_values:
            drawdown_chart_base64 = generate_chart_for_pdf(chart_labels, chart_drawdown_values, "Drawdown (%)", "Drawdown (%)", "drawdown", is_drawdown=True)

        # (선택) 상세 데이터 테이블 HTML 준비 (PDF용으로 간소화)
        # stock_data_html_for_pdf = stock_data_df_for_report.tail(10).to_html(classes='kpi-table', border=0, escape=False) if not stock_data_df_for_report.empty else "상세 데이터 없음"


        # 4. HTML 템플릿 렌더링
        rendered_html = render_template(
            'pdf_report_template.html', 
            result_data=kpis_for_report,
            equity_chart_path=equity_chart_base64, # base64 문자열 전달
            drawdown_chart_path=drawdown_chart_base64,
            # stock_data_html_for_pdf=stock_data_html_for_pdf,
            generation_date=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            current_year=datetime.now().year
        )

        # 5. WeasyPrint로 PDF 생성
        # 폰트 설정 (필요한 경우, 시스템에 설치된 폰트 경로 확인 필요)
        # font_config = FontConfiguration()
        # css = CSS(string='@font-face { font-family: NanumGothic; src: url(/path/to/NanumGothic.ttf); }', font_config=font_config) # 예시
        # html_obj = HTML(string=rendered_html, base_url=request.url_root) # base_url은 상대 경로 이미지/CSS 로드 시 필요
        # pdf_file = html_obj.write_pdf(stylesheets=[css] if 'css' in locals() else None, font_config=font_config)
        
        # 간단한 버전 (시스템 폰트에 의존)
        html_obj = HTML(string=rendered_html, base_url=request.url_root)
        pdf_file = html_obj.write_pdf()

        # 6. PDF 파일 응답 생성
        response = make_response(pdf_file)
        response.headers['Content-Type'] = 'application/pdf'
        report_filename_base = params['strategy_name']
        if is_portfolio:
            report_filename_base += "_" + "_".join(params['tickers'][:2]) # 너무 길어지지 않도록 일부 티커만 사용
        else:
            report_filename_base += "_" + params['ticker']
            
        report_filename = f"backtest_report_{report_filename_base}_{result_id}.pdf"
        report_filename = "".join(c if c.isalnum() or c in ['_', '.'] else '_' for c in report_filename) # 파일명 안전하게 처리
        response.headers['Content-Disposition'] = f'attachment; filename="{report_filename}"'
        
        return response

    except ValueError as ve:
        flash(f"PDF 리포트 생성 중 데이터 변환 오류: {str(ve)}", "danger")
    except mysql.connector.Error as e:
        print(f"DB 오류 (download_pdf): {e}"); flash(f"DB 오류: {e}", "danger")
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"일반 오류 (download_pdf): {type(e).__name__}: {e}"); flash(f"일반 오류: {type(e).__name__} - {e}", "danger")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()
    
    return redirect(url_for('view_result', result_id=result_id) if result_id else url_for('list_results'))

@app.route('/logout')
@login_required # 로그인이 되어 있어야 로그아웃 가능
def logout():
    logout_user() # Flask-Login을 통해 사용자 로그아웃 처리
    flash("성공적으로 로그아웃되었습니다.", "info")
    return redirect(url_for('home'))

if __name__ == '__main__':
    app.run(debug=True)
